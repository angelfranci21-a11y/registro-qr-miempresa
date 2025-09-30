from flask import Flask, request, render_template_string, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

# --- Configuración ---
EXCEL_FILE = "datos_qr.xlsx"
NOMBRE_HOJA_DESTINO = "Hoja 2"

# NOTA: Se eliminó el bloque de inicialización de Excel de la raíz para evitar fallos al inicio del servidor en Render.

# *******************************************************************
# RUTA 1: CARGA EL FORMULARIO (Método GET - Se ejecuta al escanear el QR)
# *******************************************************************
@app.route('/cargar_formulario', methods=['GET'])
def cargar_formulario():
    # Obtener los datos del QR (desde la URL)
    orden = request.args.get('orden', '')
    codigo = request.args.get('codigo', '')
    descripcion = request.args.get('descripcion', '') 
    lote = request.args.get('lote', '')             
    fecha_ini = request.args.get('fecha_ini', '')
    cantidad = request.args.get('cantidad', '')
    fecha_fin = request.args.get('fecha_fin', '')
    
    # Formulario HTML
    formulario_html = f"""
    <html>
    <head>
        <title>PROTOTIPO QR</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
            body {{ font-family: sans-serif; padding: 20px; }}
            input, button {{ width: 100%; padding: 10px; margin: 8px 0; box-sizing: border-box; }}
            .read-only {{ background-color: #f0f0f0; }}
            .read-only-red {{ background-color: #ffcccc; }}
        </style>
    </head>
    <body>
        <h2> NOTIFICADOR ORDENES (Confirmación)</h2>
        <form method="POST" action="/guardar_datos_final">
            
            <label>1. No. Orden:</label>
            <input type="text" name="orden" value="{orden}" class="read-only-red" readonly>
            
            <label>2. Código:</label>
            <input type="text" name="codigo" value="{codigo}" class="read-only-red" readonly>
            
            <label>3. Descripción:</label>
            <input type="text" name="descripcion" value="{descripcion}" class="read-only" readonly>
            
            <label>4. Lote:</label>
            <input type="text" name="lote" value="{lote}" class="read-only" readonly>
            
            <hr style="border-top: 1px solid #ddd; margin: 20px 0;">
            
            <label>5. Fecha Inicial:</label>
            <input type="text" name="fecha_ini" value="{fecha_ini}" class="read-only" readonly>
            
            <label style="font-weight: bold; color: darkgreen;">6. Cantidad Programada (editable):</label>
            <input type="number" name="cantidad" value="{cantidad}">
            
            <label style="font-weight: bold; color: darkgreen;">7. Fecha Final (editable):</label>
            <input type="date" name="fecha_fin" value="{fecha_fin}">
            
            <button type="submit" style="background-color: #007bff; color: white; padding: 12px;"> PULSE PARA GUARDAR EN DATA BASE</button>
        </form>
        <p><small>RECUERDA QUE TODO ESTE BIEN.</small></p>
    </body>
    </html>
    """
    return render_template_string(formulario_html)


# *******************************************************************
# RUTA 2: GUARDA LOS DATOS FINALES (Método POST - Guarda en Excel)
# *******************************************************************
@app.route('/guardar_datos_final', methods=['POST'])
def guardar_datos_final():
    try:
        # Obtener los datos del formulario
        orden = request.form['orden']
        codigo = request.form['codigo']
        descripcion = request.form['descripcion'] 
        lote = request.form['lote']             
        fecha_ini = request.form['fecha_ini']
        cantidad = request.form['cantidad']
        fecha_fin = request.form['fecha_fin']

        # Validación básica
        if not all([orden, codigo, descripcion, lote, fecha_ini, cantidad, fecha_fin]):
            return render_template_string("<h1>❌ Error: Datos incompletos al guardar.</h1>"), 400

        # --- LÓGICA DE CARGA/CREACIÓN SEGURA PARA PRODUCCIÓN ---
        if not os.path.exists(EXCEL_FILE):
             # Si el archivo NO existe (Render lo eliminó), lo creamos desde cero
             wb = openpyxl.Workbook()
             if 'Sheet' in wb.sheetnames:
                 del wb['Sheet']
             ws = wb.create_sheet(NOMBRE_HOJA_DESTINO)
             # Escribimos los encabezados
             ws.append(["No. Orden", "Código", "Descripción", "Lote", "Fecha Inicial", "Cantidad Programada", "Fecha Final"])
             wb.save(EXCEL_FILE) # Guardamos el archivo vacío con encabezados
             
        # Cargar el archivo ahora que sabemos que existe
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Seleccionar la hoja
        if NOMBRE_HOJA_DESTINO not in wb.sheetnames:
            ws = wb.create_sheet(NOMBRE_HOJA_DESTINO)
            ws.append(["No. Orden", "Código", "Descripción", "Lote", "Fecha Inicial", "Cantidad Programada", "Fecha Final"])
        else:
            ws = wb[NOMBRE_HOJA_DESTINO]
            
        # 1. Encontrar la próxima fila vacía.
        next_row = ws.max_row + 1
        
        # 2. Los datos a guardar
        datos_a_guardar = [orden, codigo, descripcion, lote, fecha_ini, cantidad, fecha_fin]
        
        # 3. Escribir celda por celda (garantiza el inicio en Columna A)
        for col_idx, valor in enumerate(datos_a_guardar, 1):
            ws.cell(row=next_row, column=col_idx, value=valor)
            
        wb.save(EXCEL_FILE)

        # Respuesta de éxito
        return render_template_string(f"""
        <html>
        <body style="font-family: sans-serif; text-align: center; background-color: #ccffcc;">
            <h1 style="color: green;">✅ ¡Datos Guardados!</h1>
            <p>Orden <strong>{orden}</strong> y Lote <strong>{lote}</strong> registrados correctamente en la Fila {next_row}.</p>
        </body>
        </html>
        """), 200

    except Exception as e:
        # Mostrar el error si ocurre uno (útil para debug)
        return render_template_string(f"<h1>❌ Error al guardar: {str(e)}</h1>"), 500

# NOTA: Se eliminó la sección "if __name__ == '__main__':" para que funcione con Gunicorn/Render.
